from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, JsonResponse
import os
import io
from datetime import datetime


def convert_txt_files(files):
    current_dir = os.path.dirname(__file__)
    template_path = os.path.normpath(
        os.path.join(current_dir, "..", "file_templates", "Шаблон_excel51_to_txt.xlsx")
    )

    wb = load_workbook(template_path)
    ws = wb.active

    row = ws.max_row
    rows_converted = 0

    mapping = {
        "Номер": "C", "Дата": "D", "Сумма": "E",
        "ДатаСписано": "F", "ДатаПоступило": "G",
        "Плательщик": "H", "Плательщик1": "H",
        "ПлательщикИНН": "I", "ПлательщикРасчСчет": "J",
        "ПлательщикБанк1": "K", "ПлательщикБанк2": "K",
        "ПолучательСчет": "L",
        "Получатель": "M", "Получатель1": "M",
        "ПолучательИНН": "N", "ПолучательРасчСчет": "O",
        "ПолучательБанк1": "P", "ПолучательБанк2": "P",
        "НазначениеПлатежа": "Q",
    }

    current_data = {}
    post = True

    for file in files:
        content = file.read()

        if content.startswith(b'\xef\xbb\xbf'):
            content = content[3:]

        text = content.decode("cp1251")
        lines = text.splitlines()

        for line in lines:
            if line.startswith("СекцияДокумент="):
                if current_data:
                    row += 1
                    write_row(ws, row, current_data, mapping, post)
                    rows_converted += 1
                    current_data = {}
                    post = True
                current_data["A"] = line.split("=", 1)[1]

            elif line.strip() == "КонецДокумента":
                row += 1
                write_row(ws, row, current_data, mapping, post)
                rows_converted += 1
                current_data = {}
                post = True

            elif "=" in line:
                k, v = line.split("=", 1)
                if k in mapping:
                    current_data[mapping[k]] = v
                if k == "ДатаСписано" and v:
                    post = False

    if current_data:
        row += 1
        write_row(ws, row, current_data, mapping, post)
        rows_converted += 1

    if rows_converted == 0:
        return JsonResponse({
            'success': False,
            'error': 'Файлы не содержат данных для конвертации. Проверьте формат .txt'
        }, status=400)

    for r in ws.iter_rows():
        for cell in r:
            cell.alignment = Alignment(vertical="center", horizontal="left")

    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    today_str = datetime.now().strftime("%Y%m%d")
    filename = f"51_converted_{today_str}.xlsx"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


def write_row(ws, row, data, mapping, post):
    if "A" in data:
        ws[f"A{row}"] = data["A"]

    for col, value in data.items():
        if col != "A":
            ws[f"{col}{row}"] = value

    ws[f"B{row}"] = "Списание" if not post else "Поступление"